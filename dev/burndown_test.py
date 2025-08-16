import os, zipfile
from pathlib import Path
from datetime import datetime
from PIL import Image, ImageDraw
from docx import Document
from pypdf import PdfWriter
import csv

# 📁 Destination folder
OUT_DIR = Path("meta_dirty_bundle")
OUT_DIR.mkdir(exist_ok=True)

def create_txt():
    path = OUT_DIR / "dirty.txt"
    with open(path, "w") as f:
        f.write("Name: John Doe\nEmail: john@example.com\nPhone: 123-456-7890\n")
    os.utime(path, (1609459200, 1609459200))  # Jan 1, 2021

def create_csv():
    path = OUT_DIR / "dirty.csv"
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["name", "email", "ssn"])
        writer.writerow(["Jane Roe", "jane@example.com", "123-45-6789"])

def create_docx():
    path = OUT_DIR / "dirty.docx"
    doc = Document()
    doc.add_paragraph("Author: MetaWriter\nSSN: 000-00-0000")
    doc.save(path) # type: ignore

def create_pdf():
    path = OUT_DIR / "dirty.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    writer.add_metadata({
        "/Author": "MetaMaker",
        "/Title": "Dirty PDF",
        "/Subject": "Test PII"
    })
    with open(path, "wb") as f:
        writer.write(f)

def create_xlsx():
    from openpyxl import Workbook
    path = OUT_DIR / "dirty.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MetaSheet" # type: ignore
    ws['A1'] = "Name" # type: ignore
    ws['B1'] = "Email" # type: ignore
    ws['A2'] = "Rick" # type: ignore
    ws['B2'] = "rick@example.com" # type: ignore
    wb.save(path)

def create_jpg():
    path = OUT_DIR / "dirty.jpg"
    img = Image.new("RGB", (100, 100), color="red")
    draw = ImageDraw.Draw(img)
    draw.text((10, 40), "Meta", fill="white")
    exif_bytes = b"Exif\x00\x00" + b"\xff\xd8" * 10  # dummy EXIF
    img.save(path, exif=exif_bytes)

def create_heic():
    pass

def zip_payload():
    zip_path = "meta_test_payload.zip"
    with zipfile.ZipFile(zip_path, "w") as zipf:
        for file in OUT_DIR.iterdir():
            zipf.write(file, arcname=file.name)
    print(f"✅ ZIP bundle created: {zip_path}")

if __name__ == "__main__":
    create_txt()
    create_csv()
    create_docx()
    create_pdf()
    create_xlsx()
    create_jpg()
    # create_heic()  # Optional: requires external tools
    zip_payload()
